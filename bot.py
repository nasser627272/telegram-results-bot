import logging
import os
from telegram import Update
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
)
import openpyxl
from openpyxl import load_workbook

# تكوين التسجيل
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# مسار ملف Excel (يمكن تغييره عبر متغير البيئة)
EXCEL_FILE = os.getenv('EXCEL_FILE_PATH', 'results.xlsx')

class ExcelManager:
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance.init_excel()
        return cls._instance
    
    def init_excel(self):
        try:
            self.wb = load_workbook(EXCEL_FILE)
            self.sheet = self.wb.active
            self.results_data = self.load_data()
            logger.info(f"تم تحميل {len(self.results_data)} سجل من ملف Excel")
        except Exception as e:
            logger.error(f"خطأ في تحميل ملف Excel: {e}")
            raise
    
    def load_data(self):
        """تحميل البيانات من Excel إلى قاموس"""
        data = {}
        for row in self.sheet.iter_rows(values_only=True):
            if row and row[0]:
                data[str(row[0])] = {
                    'name': row[1] if len(row) > 1 else 'غير متوفر',
                    'result': row[2] if len(row) > 2 else 'غير متوفر'
                }
        return data
    
    def get_student_data(self, seating_no):
        """البحث عن بيانات الطالب"""
        return self.results_data.get(seating_no)

# إنشاء مدير Excel
try:
    excel_manager = ExcelManager()
except Exception as e:
    logger.error(f"فشل تهيئة مدير Excel: {e}")
    exit(1)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """معالجة أمر /start"""
    user = update.effective_user
    welcome_msg = (
        f"مرحباً {user.mention_html()} 👋\n\n"
        "أنا بوت البحث عن نتائج الطلاب.\n"
        "أرسل لي <b>رقم الجلوس</b> وسأعرض لك النتيجة."
    )
    await update.message.reply_html(welcome_msg)

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """معالجة الرسائل النصية"""
    seating_no = update.message.text.strip()
    
    if not seating_no:
        await update.message.reply_text("⚠️ الرجاء إدخال رقم الجلوس")
        return
    
    student_data = excel_manager.get_student_data(seating_no)
    
    if student_data:
        result_msg = (
            "🎓 <b>بيانات الطالب:</b>\n\n"
            f"<b>رقم الجلوس:</b> {seating_no}\n"
            f"<b>الاسم:</b> {student_data['name']}\n"
            f"<b>النتيجة:</b> {student_data['result']}"
        )
    else:
        result_msg = f"⚠️ لا توجد نتائج لرقم الجلوس {seating_no}"
    
    await update.message.reply_html(result_msg)

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """معالجة الأخطاء"""
    logger.error(f"حدث خطأ: {context.error}", exc_info=True)
    
    if update and isinstance(update, Update):
        if update.message:
            await update.message.reply_text(
                "❌ حدث خطأ غير متوقع. الرجاء المحاولة لاحقاً."
            )

def main() -> None:
    """تشغيل البوت"""
    try:
        # الحصول على توكن البوت من متغير البيئة
        bot_token = os.getenv('TELEGRAM_BOT_TOKEN')
        if not bot_token:
            raise ValueError("لم يتم تعيين TELEGRAM_BOT_TOKEN في متغيرات البيئة")
        
        # بناء التطبيق
        app = ApplicationBuilder().token(bot_token).build()
        
        # تسجيل المعالجات
        app.add_handler(CommandHandler("start", start))
        app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
        
        # معالج الأخطاء
        app.add_error_handler(error_handler)
        
        # معلومات البدء
        logger.info("Starting bot...")
        print("""
        *********************************
        *       تم تشغيل البوت بنجاح       *
        *********************************
        """)
        
        # بدء البوت
        app.run_polling()
    
    except Exception as e:
        logger.error(f"فشل تشغيل البوت: {e}")
        exit(1)

if __name__ == "__main__":
    main()
